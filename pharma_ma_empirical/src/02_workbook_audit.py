"""Phase A: complete workbook structure audit for both DiD workbooks."""
import json
import openpyxl
import pandas as pd
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
AUDIT = ROOT / "outputs" / "audit"
AUDIT.mkdir(parents=True, exist_ok=True)

FILES = {
    "treatment": RAW / "Treatment_Group_DiD_Data_Workbook.xlsx",
    "control": RAW / "Control_Group_DiD_Data_Workbook.xlsx",
}

ERR_TOKENS = ("#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", "#NUM!")

structure_rows = []
formula_err_rows = []
sample_dump = {}

for label, path in FILES.items():
    wb_f = openpyxl.load_workbook(path, data_only=False)   # formulas
    wb_v = openpyxl.load_workbook(path, data_only=True)    # cached values
    for ws in wb_f.worksheets:
        wsv = wb_v[ws.title]
        n_formulas = 0
        n_err_formula = 0
        n_err_cached = 0
        first_rows = []
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            if r_idx <= 3:
                first_rows.append([str(c.value)[:40] if c.value is not None else "" for c in row[:15]])
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    n_formulas += 1
                    if any(t in c.value for t in ERR_TOKENS):
                        n_err_formula += 1
                        if n_err_formula <= 5 or True:
                            formula_err_rows.append({
                                "workbook": label, "sheet": ws.title, "cell": c.coordinate,
                                "kind": "formula_contains_error_ref", "content": c.value[:120]})
                v = wsv[c.coordinate].value
                if isinstance(v, str) and v in ERR_TOKENS:
                    n_err_cached += 1
                    formula_err_rows.append({
                        "workbook": label, "sheet": ws.title, "cell": c.coordinate,
                        "kind": "cached_error_value", "content": v})
        structure_rows.append({
            "workbook": label, "sheet": ws.title,
            "max_row": ws.max_row, "max_col": ws.max_column,
            "n_formulas": n_formulas,
            "n_formula_err": n_err_formula, "n_cached_err": n_err_cached,
            "n_merged_ranges": len(ws.merged_cells.ranges),
            "visible": ws.sheet_state,
        })
        sample_dump[f"{label}::{ws.title}"] = first_rows
    wb_f.close(); wb_v.close()

pd.DataFrame(structure_rows).to_excel(AUDIT / "workbook_structure.xlsx", index=False)
err_df = pd.DataFrame(formula_err_rows)
# summarise errors to keep file manageable
if len(err_df):
    err_summary = err_df.groupby(["workbook", "sheet", "kind"]).size().reset_index(name="n_cells")
else:
    err_summary = pd.DataFrame()
with pd.ExcelWriter(AUDIT / "formula_error_audit.xlsx") as xw:
    err_summary.to_excel(xw, sheet_name="summary", index=False)
    err_df.head(5000).to_excel(xw, sheet_name="detail_first5000", index=False)

with open(AUDIT / "sheet_first_rows.json", "w") as f:
    json.dump(sample_dump, f, indent=1)

print(pd.DataFrame(structure_rows).to_string())
print("\nERROR SUMMARY:")
print(err_summary.to_string() if len(err_summary) else "none")
